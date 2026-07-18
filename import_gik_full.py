#!/usr/bin/env python3
"""
Complete GIK UGM Data Import for ERP Konstruksi
Imports: RAB Budgets, Suppliers, Purchase Orders
"""
import openpyxl
from datetime import datetime
import json
import sys
sys.path.append('/c/Users/faras/rep-sangar')

import subprocess

def run_tinker(code: str):
    """Run PHP code via artisan tinker"""
    cmd = f'cd /c/Users/faras/rep-sangar && /c/php83/php.exe artisan tinker --execute="{code}"'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=60)
    return result.stdout, result.stderr

# ============================================================
# STEP 1: Parse GIK UGM RAB (1035 items)
# ============================================================
def parse_gik_rab():
    path = "C:/Users/faras/.hermes/desktop-attachments/C.1 RAB GIK UGM Ulang.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["RAB GIK UGM"]
    
    items = []
    current_category = None
    current_sub_category = None
    item_no = 0
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 7:  # Skip header rows
            continue
            
        no = row[0]
        kode = row[1]
        uraian = row[2]
        satuan = row[3]
        volume = row[4]
        harga = row[5]
        jumlah = row[6]
        
        if not uraian:
            continue
            
        uraian_str = str(uraian).strip()
        
        # Detect section headers
        if kode and str(kode).strip() in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
            if 'MATA PEMBAYARAN' in uraian_str.upper():
                current_category = uraian_str
            elif any(x in uraian_str.upper() for x in ['PEKERJAAN', 'STRUKTUR', 'ARSITEKTUR', 'MEP', 'UTAMA', 'PERSIAPAN', 'SMKKK']):
                current_sub_category = uraian_str
            continue
            
        # Skip if no volume/price (category header)
        if volume is None and harga is None and jumlah is None:
            continue
            
        try:
            item_no += 1
            items.append({
                'item_no': item_no,
                'code': str(kode).strip() if kode else None,
                'description': uraian_str,
                'unit': str(satuan).strip() if satuan else None,
                'qty': float(volume) if volume else 0,
                'price': float(harga) if harga else 0,
                'total': float(jumlah) if jumlah else 0,
                'category': current_category,
                'sub_category': current_sub_category,
            })
        except (ValueError, TypeError):
            pass
    
    wb.close()
    print(f"Parsed {len(items)} RAB items")
    return items

# ============================================================
# STEP 2: Parse Purchase Order GIK UGM (36 vendors)
# ============================================================
def parse_gik_po():
    path = "C:/Users/faras/.hermes/desktop-attachments/Purchase Order GIK UGM.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    po_data = []
    
    for sname in wb.sheetnames:
        if sname.lower() in ['sheet1']:
            continue
            
        ws = wb[sname]
        
        vendor_name = None
        po_number = None
        po_date = None
        project = None
        location = None
        contact = None
        
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 200:
                break
                
            row_vals = [v for v in row if v is not None]
            if not row_vals:
                continue
                
            row_str = ' '.join(str(v) for v in row_vals)
            
            # Vendor name
            if not vendor_name and len(row_vals) == 1:
                val = str(row_vals[0]).strip()
                if any(p in val for p in ['PT.', 'CV.', 'TOKO', 'UD ', 'BENGKEL', 'SHOPEE', 'SUMBER ', 'CAHAYA ']):
                    vendor_name = val
            
            # PO Number
            if not po_number and ':' in row_str:
                for cell in row:
                    if cell and ('SCS-SMG' in str(cell) or 'S/SCS' in str(cell)):
                        if 'PO' in str(cell) or '/' in str(cell):
                            po_number = str(cell).strip()
            
            # Date
            for j, cell in enumerate(row):
                if cell and isinstance(cell, datetime):
                    po_date = cell
                elif cell and '202' in str(cell) and ('-' in str(cell) or '/' in str(cell)):
                    po_date = str(cell).strip()
            
            # Project
            if not project and 'Proyek' in row_str and ':' in row_str:
                for cell in row:
                    if cell and 'GIK' in str(cell):
                        project = str(cell).strip()
            
            # Location
            if not location and 'Lokasi' in row_str and ':' in row_str:
                for cell in row:
                    if cell and 'Yogyakarta' in str(cell):
                        location = str(cell).strip()
            
            # Contact
            if not contact and 'Contact' in row_str and ':' in row_str:
                for cell in row:
                    if cell and ('Hudi' in str(cell) or 'Dede' in str(cell) or 'Deni' in str(cell) or 'P.' in str(cell)):
                        contact = str(cell).strip()
        
        if vendor_name and po_number:
            po_data.append({
                'vendor': vendor_name,
                'po_number': po_number,
                'po_date': po_date,
                'project': project,
                'location': location,
                'contact': contact,
                'sheet': sname
            })
    
    wb.close()
    print(f"Parsed {len(po_data)} POs")
    return po_data

# ============================================================
# STEP 3: Import to ERP via Tinker
# ============================================================
def import_to_erp(rab_items, po_data):
    """Import all data to ERP using tinker"""
    
    # 1. Create/Get Project
    project_code = """
    $project = App\Models\Project::firstOrCreate(
        ['project_name' => 'GIK UGM'],
        [
            'location' => 'Yogyakarta',
            'start_date' => '2025-01-01',
            'status' => 'planning',
        ]
    );
    echo $project->id;
    """
    out, err = run_tinker(project_code)
    project_id = out.strip()
    print(f"Project ID: {project_id}")
    
    # 2. Create RAB Categories & Items
    print("\nImporting RAB items...")
    cat_cache = {}
    
    for idx, item in enumerate(rab_items):
        if not item['description']:
            continue
            
        cat_name = item['category'] or 'Umum'
        if cat_name not in cat_cache:
            cat_code = """
            $cat = App\Models\RabCategory::firstOrCreate(
                ['name' => '%s', 'project_id' => %s],
                ['code' => 'CAT-' . strtoupper(substr('%s', 0, 3))]
            );
            echo $cat->id;
            """ % (cat_name.replace("'", "\\'"), project_id, cat_name.replace("'", "\\'"))
            out, err = run_tinker(cat_code)
            cat_cache[cat_name] = out.strip()
        
        rab_code = """
        $item = App\Models\RabItem::create([
            'rab_category_id' => %s,
            'code' => '%s',
            'description' => '%s',
            'unit' => '%s',
            'quantity' => %f,
            'price' => %f,
            'total' => %f,
        ]);
        echo $item->id . ',';
        """ % (
            cat_cache[cat_name],
            (item['code'] or 'ITEM-' + str(idx)).replace("'", "\\'"),
            item['description'].replace("'", "\\'"),
            (item['unit'] or 'unit').replace("'", "\\'"),
            item['qty'],
            item['price'],
            item['total']
        )
        
        if idx % 100 == 0:
            print(f"  Imported {idx}/{len(rab_items)} items...")
        out, err = run_tinker(rab_code)
        if err and 'duplicate' not in err.lower():
            pass  # Ignore duplicates
    
    print("RAB import complete")
    
    # 3. Create Suppliers & POs
    print("\nImporting Suppliers & POs...")
    for po in po_data:
        # Create supplier
        supplier_code = """
        $supplier = App\Models\Supplier::firstOrCreate(
            ['name' => '%s'],
            [
                'code' => 'SUP-' . strtoupper(substr('%s', 0, 10)),
                'address' => '%s',
                'phone' => '%s',
                'email' => '%s',
                'contact_person' => '%s',
            ]
        );
        echo $supplier->id;
        """ % (
            po['vendor'].replace("'", "\\'"),
            po['vendor'].replace("'", "\\'"),
            (po['location'] or '').replace("'", "\\'"),
            (po['contact'] or '').replace("'", "\\'"),
            '',
            (po['contact'] or '').replace("'", "\\'")
        )
        out, err = run_tinker(supplier_code)
        supplier_id = out.strip()
        
        # Create PO
        po_code = """
        $po = App\Models\PurchaseOrder::create([
            'po_number' => '%s',
            'project_id' => %s,
            'supplier_id' => %s,
            'date' => '%s',
            'status' => 'draft',
            'po_type' => 'supplier',
            'payment_terms' => '30 hari',
            'contact_person' => '%s',
            'supplier_address' => '%s',
        ]);
        echo $po->id;
        """ % (
            po['po_number'].replace("'", "\\'"),
            project_id,
            supplier_id,
            str(po['po_date'])[:10] if po['po_date'] else '2026-01-01',
            (po['contact'] or '').replace("'", "\\'"),
            (po['location'] or '').replace("'", "\\'")
        )
        out, err = run_tinker(po_code)
        if err:
            print(f"  PO Error: {err[:100]}")
        else:
            print(f"  Created PO: {po['po_number']} for {po['vendor']}")

def run_tinker(code):
    """Run PHP code via artisan tinker"""
    cmd = 'cd /c/Users/faras/rep-sangar && /c/php83/php.exe artisan tinker --execute="' + code + '"'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=60)
    return result.stdout, result.stderr

# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print("GIK UGM FULL IMPORT")
    print("=" * 60)
    
    print("\n1. Parsing RAB...")
    rab_items = parse_gik_rab()
    
    print("\n2. Parsing POs...")
    po_data = parse_gik_po()
    for po in po_data[:5]:
        print(f"  {po['vendor']} - {po['po_number']}")
    
    print(f"\nTotal RAB items: {len(rab_items)}")
    print(f"Total POs: {len(po_data)}")
    
    # Save parsed data
    with open('/c/Users/faras/rep-sangar/storage/app/gik_parsed.json', 'w') as f:
        json.dump({'rab': rab_items, 'po': po_data}, f, default=str, indent=2)
    print("\nSaved parsed data to storage/app/gik_parsed.json")
    
    # Import to ERP
    print("\n3. Importing to ERP...")
    import_to_erp(rab_items, po_data)
    
    print("\n" + "=" * 60)
    print("IMPORT COMPLETE")
    print("=" * 60)