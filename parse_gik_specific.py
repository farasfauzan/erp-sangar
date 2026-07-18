import openpyxl
import json
from pathlib import Path

def parse_gik_ugm_rab():
    """Parse the main RAB GIK UGM sheet with 1302 rows"""
    path = "C:/Users/faras/.hermes/desktop-attachments/C.1 RAB GIK UGM Ulang.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    # The main RAB sheet
    ws = wb["RAB GIK UGM"]
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    # Print header rows to understand structure
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 15:
            print(f"R{i+1}: {row}")
        else:
            break
    
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 7:  # Skip header rows
            continue
        
        # Expected columns: No, Kode/Jenis Barang, Satuan, Volume, Harga Satuan, Jumlah Harga, ...
        # R8: (1, None, 2, 3, 4, 5, 6, None, None, None, None, None, None, None, None)
        # R9: ('A.', None, 'MATA PEMBAYARAN UMUM', None, None, None, None, None, None, None, None, None, None, None, None)
        # R10: (None, None, 'PEKERJAAN PERSIAPAN', None, None, None, None, None, None, None, None, None, None, None, None)
        # R11: (None, None, 'Pengukuran dan pemasangan Bouwplank', 'm1', 120, 62700, 7524000, None, None, None, None, None, None, None, None)
        
        # row[0] = No (can be None for sub-items)
        # row[1] = Kode/Section (A, B, C, or None)
        # row[2] = Uraian/Description
        # row[3] = Satuan
        # row[4] = Volume
        # row[5] = Harga Satuan
        # row[6] = Jumlah Harga
        
        uraian = row[2] if len(row) > 2 else None
        satuan = row[3] if len(row) > 3 else None
        volume = row[4] if len(row) > 4 else None
        harga_satuan = row[5] if len(row) > 5 else None
        jumlah_harga = row[6] if len(row) > 6 else None
        
        # Skip if no uraian or it's a section header
        if not uraian or uraian in ['MATA PEMBAYARAN UMUM', 'MATA PEMBAYARAN PERKIRAAN BIAYA PENERAPAN SMKKK']:
            continue
        
        # Skip if it's a category header (no volume/price)
        if volume is None and harga_satuan is None and jumlah_harga is None:
            continue
        
        try:
            items.append({
                "code": row[1] if len(row) > 1 else None,
                "description": str(uraian).strip(),
                "unit": str(satuan).strip() if satuan else None,
                "qty": float(volume) if volume else 0,
                "price": float(harga_satuan) if harga_satuan else 0,
                "total": float(jumlah_harga) if jumlah_harga else 0,
                "category": None  # Will fill from section
            })
        except (ValueError, TypeError):
            pass
    
    print(f"Parsed {len(items)} items from RAB GIK UGM")
    
    # Now parse category from section headers
    current_category = None
    for item in items:
        if item['code'] in ['A', 'B', 'C', 'D', 'E']:
            current_category = item['description']
            item['category'] = current_category
        else:
            item['category'] = current_category
    
    # Filter out section headers
    final_items = [i for i in items if i['code'] not in ['A', 'B', 'C', 'D', 'E']]
    
    print(f"Final items: {len(final_items)}")
    for i in final_items[:10]:
        print(f"  {i['code']} | {i['description'][:50]} | {i['unit']} | {i['qty']} | {i['price']:,.0f} | {i['total']:,.0f} | {i['category']}")
    
    wb.close()
    return final_items

def parse_gik_ugm_po():
    """Parse Purchase Order GIK UGM"""
    path = "C:/Users/faras/.hermes/desktop-attachments/Purchase Order GIK UGM.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"PO Sheets: {wb.sheetnames}")
    
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\nSheet: {sname} | Rows: {ws.max_row} | Cols: {ws.max_column}")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 20:
                print(f"  R{i+1}: {row}")
            else:
                break
    wb.close()

if __name__ == "__main__":
    print("=" * 60)
    print("PARSING GIK UGM RAB")
    print("=" * 60)
    items = parse_gik_ugm_rab()
    
    print("\n" + "=" * 60)
    print("PARSING GIK UGM PO")
    print("=" * 60)
    parse_gik_ugm_po()