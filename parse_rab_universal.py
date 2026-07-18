#!/usr/bin/env python3
"""
Universal RAB Parser for PT. SINAR CERAH SEMPURNA
Handles multiple Excel formats from different vendors/projects
"""
import openpyxl
import json
import re
from pathlib import Path
from typing import List, Dict, Any, Optional

class RABParser:
    def __init__(self):
        self.standard_columns = ['code', 'description', 'category', 'unit', 'qty', 'price', 'total', 'source_sheet', 'source_file']
    
    def detect_header_row(self, ws, max_check=20) -> int:
        """Find the row containing column headers"""
        keywords = ['nomor', 'no.', 'uraian', 'jenis', 'satuan', 'volume', 'harga', 'jumlah', 'kode', 'item', 'deskripsi']
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_check, values_only=True)):
            row_str = ' '.join(str(c).lower() for c in row if c)
            if any(kw in row_str for kw in keywords):
                return i + 1  # 1-indexed
        return 1
    
    def normalize_item(self, raw: Dict, category: str = '', source_sheet: str = '', source_file: str = '') -> Dict:
        """Map raw row to standard schema"""
        # Try multiple possible key variations
        code = raw.get('nomor') or raw.get('no.') or raw.get('no') or raw.get('kode') or raw.get('kode item') or raw.get('code') or ''
        desc = raw.get('uraian') or raw.get('uraian pekerjaan') or raw.get('jenis barang/jasa') or raw.get('deskripsi') or raw.get('nama barang') or raw.get('description') or ''
        unit = raw.get('satuan') or raw.get('satuan unit') or raw.get('unit') or ''
        qty = raw.get('volume') or raw.get('qty') or raw.get('jumlah') or raw.get('kuantitas') or raw.get('kuantitas') or 0
        price = raw.get('harga satuan') or raw.get('harga satuan (rp)') or raw.get('harga') or raw.get('harga_satuan') or 0
        total = raw.get('jumlah harga') or raw.get('jumlah harga satuan') or raw.get('total') or raw.get('total (rp.)') or 0
        
        # Calculate total if missing
        try:
            if (not total or total == 0) and qty and price:
                total = float(qty) * float(price)
        except:
            pass
        
        return {
            'code': str(code).strip() if code else '',
            'description': str(desc).strip() if desc else '',
            'category': str(category).strip(),
            'unit': str(unit).strip() if unit else '',
            'qty': float(qty) if qty else 0,
            'price': float(price) if price else 0,
            'total': float(total) if total else 0,
            'source_sheet': source_sheet,
            'source_file': source_file
        }
    
    def parse_sheet(self, ws, source_file: str) -> List[Dict]:
        """Parse a single worksheet"""
        items = []
        header_row = self.detect_header_row(ws)
        
        # Get headers
        headers = []
        for cell in ws[header_row]:
            val = cell.value
            headers.append(str(val).strip().lower() if val else f'col_{len(headers)}')
        
        current_category = ''
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1, values_only=True), start=header_row+1):
            if not any(v for v in row if v is not None):
                continue
            
            # Build row dict
            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            
            # Check if it's a category header (has description but no qty/price)
            first_val = row[0]
            has_qty = any(isinstance(v, (int, float)) and v > 0 for v in row[3:7])  # rough check
            has_price = any(isinstance(v, (int, float)) and v > 1000 for v in row[5:7])
            
            # Category detection: row with text in col 2/3 but no qty/price
            desc_col = row[2] if len(row) > 2 else row[1] if len(row) > 1 else None
            if desc_col and not has_qty and not has_price and str(desc_col).strip():
                # Likely a category header
                text = str(desc_col).strip().upper()
                if any(kw in text for kw in ['MATA PEMBAYARAN', 'PEKERJAAN', 'STRUKTUR', 'ARSITEKTUR', 'MEP', 'PERSIAPAN', 'SMKKK']):
                    current_category = str(desc_col).strip()
                    continue
            
            # Data row: has numeric qty or price
            code_val = row[0] if len(row) > 0 else None
            if isinstance(code_val, (int, float)) or (isinstance(code_val, str) and code_val.strip().isdigit()):
                # This looks like a data row
                item = self.normalize_item(row_dict, current_category, ws.title, source_file)
                if item['description'] or item['qty'] or item['price']:
                    items.append(item)
        
        return items
    
    def parse_file(self, filepath: str) -> List[Dict]:
        """Parse entire Excel file"""
        all_items = []
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                items = self.parse_sheet(ws, Path(filepath).name)
                if items:
                    all_items.extend(items)
                    print(f"  {sheet_name}: {len(items)} items")
            except Exception as e:
                print(f"  {sheet_name}: ERROR - {e}")
        
        wb.close()
        return all_items

def main():
    parser = RABParser()
    base = Path("C:/Users/faras/.hermes/desktop-attachments")
    
    # All RAB-related files
    rab_files = [
        "C.1 RAB GIK UGM Ulang.xlsx",
        "GIK UGM TAHAP II.xlsx", 
        "GIK.xlsx",
        "Purchase Order GIK UGM.xlsx",
        "C.1 Rev. RAB SEKOLAH GORONTALO.xlsx",
        "Purchase Order Sekolah Rakyat Gorontalo 1.xlsx",
        "SEKOLAH RAKYAT GORONTALO.xlsx",
        "SR GORONTALO 2.xlsx",
        "C.1 RAB SPAM JINGAH.xlsx",
        "C.1 RAB WFC BARITO.xlsx",
        "B. RAB RSUD MENTAWAI FIX.xlsx",
        "Rab Rs Muara Teweh 2026.xlsx"
    ]
    
    all_data = {}
    
    for fname in rab_files:
        path = base / fname
        if not path.exists():
            print(f"SKIP: {fname} not found")
            continue
        
        print(f"\n{'='*60}")
        print(f"PARSING: {fname}")
        items = parser.parse_file(str(path))
        all_data[fname] = items
        print(f"TOTAL: {len(items)} items")
    
    # Save combined JSON
    output = Path("/c/Users/faras/rep-sangar/storage/app/rab_parsed.json")
    output.parent.mkdir(parents=True, exist_ok=True)
    with open(output, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    
    print(f"\n\nSaved to {output}")
    print(f"Files parsed: {len(all_data)}")
    for fname, items in all_data.items():
        print(f"  {fname}: {len(items)} items")

if __name__ == "__main__":
    main()