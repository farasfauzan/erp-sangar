import openpyxl

def parse_excel_file(filepath, name):
    print(f"\n{'='*60}")
    print(f"FILE: {name}")
    print(f"{'='*60}")
    
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\n  Sheet: {sname} | Rows: {ws.max_row} | Cols: {ws.max_column}")
            
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 15:
                    print(f"    R{i+1}: {row}")
                else:
                    break
        wb.close()
    except Exception as e:
        print(f"  ERROR: {e}")

base = "C:/Users/faras/.hermes/desktop-attachments/"

files = [
    ("C.1 Rev. RAB SEKOLAH GORONTALO.xlsx", "RAB Gorontalo"),
    ("Purchase Order Sekolah Rakyat Gorontalo 1.xlsx", "PO Gorontalo 1"),
    ("Purchase Order Sekolah Rakyat Gorontalo 1-2.xlsx", "PO Gorontalo 1-2"),
    ("SEKOLAH RAKYAT GORONTALO.xlsx", "Master Gorontalo"),
    ("SEKOLAH RAKYAT GORONTALO-2.xlsx", "Master Gorontalo-2"),
    ("SR GORONTALO 2.xlsx", "SR Gorontalo 2"),
    ("SR GORONTALO 2-2.xlsx", "SR Gorontalo 2-2"),
]

for fname, label in files:
    parse_excel_file(base + fname, label)