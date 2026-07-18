import openpyxl
import sys

files = [
    ("Sekolah Gorontalo", r"C:\Users\faras\.hermes\desktop-attachments\C.1 Rev. RAB SEKOLAH GORONTALO-2.xlsx"),
    ("GIK UGM", r"C:\Users\faras\.hermes\desktop-attachments\C.1 RAB GIK UGM Ulang-2.xlsx"),
    ("RSUD Mentawai", r"C:\Users\faras\.hermes\desktop-attachments\B. RAB RSUD MENTAWAI FIX-3.xlsx"),
]

for name, path in files:
    print(f"\n=== {name} ===")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for s in wb.sheetnames[:5]:
            ws = wb[s]
            if ws.max_row < 5:
                continue
            rows = list(ws.iter_rows(max_row=20, values_only=True))
            for i, r in enumerate(rows):
                vals = [str(c)[:30] if c else None for c in r[:14]]
                if any(v and ('uraian' in v.lower() or 'volume' in v.lower() or 'jumlah' in v.lower() or 'harga' in v.lower() or 'nama barang' in v.lower()) for v in vals if v):
                    print(f"Sheet: {s} (rows={ws.max_row}) | Header {i+1}: {vals}")
                    break
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
