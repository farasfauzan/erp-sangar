import openpyxl

files = [
    "C.1 RAB GIK UGM Ulang.xlsx",
    "GIK UGM TAHAP II.xlsx",
    "GIK.xlsx",
    "Purchase Order GIK UGM.xlsx",
    "Purchase Order Sekolah Rakyat Gorontalo 1.xlsx",
    "SEKOLAH RAKYAT GORONTALO.xlsx",
    "SR GORONTALO 2.xlsx"
]

import os

# Get actual files from directory
attach_dir = "C:\\Users\\faras\\.hermes\\desktop-attachments\\"
actual_files = os.listdir(attach_dir)

# Map target names to actual filenames
target_map = {
    "C.1 RAB GIK UGM Ulang.xlsx": "C.1 RAB GIK UGM Ulang.xlsx",
    "GIK UGM TAHAP II.xlsx": "GIK UGM TAHAP II.xlsx",
    "GIK.xlsx": "GIK.xlsx",
    "Purchase Order GIK UGM.xlsx": "Purchase Order GIK UGM.xlsx",
    "Purchase Order Sekolah Rakyat Gorontalo 1.xlsx": "Purchase Order Sekolah Rakyat Gorontalo 1.xlsx",
    "SEKOLAH RAKYAT GORONTALO.xlsx": "SEKOLAH RAKYAT GORONTALO.xlsx",
    "SR GORONTALO 2.xlsx": "SR GORONTALO 2.xlsx",
}

for target, actual in target_map.items():
    path = os.path.join(attach_dir, actual)
    if not os.path.exists(path):
        print(f"NOT FOUND: {target} -> {actual}")
        continue
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        print(f"\n{'='*60}")
        print(f"FILE: {target} ({actual})")
        print(f"Sheets: {wb.sheetnames}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"  Sheet: {sname} | Rows: {ws.max_row} | Cols: {ws.max_column}")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 15:
                    print(f"    R{i+1}: {row}")
                else:
                    break
        wb.close()
    except Exception as e:
        print(f"  ERROR: {e}")